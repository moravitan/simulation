import numpy as np
from openpyxl import *

excel_path = "C:\\Users\איתן אביטן\Downloads\לימודים\סימולציה לרשתות תקשורת\פרויקט\data.xlsx"


def calculate_arrival_time():
    wb = load_workbook(excel_path)
    ws = wb["2"]
    # time for hour
    arrival_time = np.random.poisson(20, 10000)

    arrival_time = list(set(arrival_time))
    arrival_time = np.sort(arrival_time)
    max_value = max(arrival_time)

    j = 0
    for i in range(6, len(arrival_time)):
        wc_ell = ws.cell(i, 2)
        wc_ell.value = round(arrival_time[j] / max_value, 3) * 3600
        j += 1

    wb.save(excel_path)


def calculate_inter_arrival_time():
    wb = load_workbook(excel_path)
    ws = wb["2"]

    wc_cell = ws.cell(6, 3)
    wc_cell.value = ws.cell(6, 2).value

    for i in range(7, 34):
        wc_cell = ws.cell(i, 3)
        wc_cell.value = ws.cell(i, 2).value - ws.cell(i - 1, 2).value

    wb.save(excel_path)


def calculate_service_time(min, median, max, col):
    wb = load_workbook(excel_path)
    ws = wb["2"]

    service_time = np.random.triangular(min, median, max, 32)
    service_time = list(set(service_time))

    j = 0
    for i in range(6, 34):
        wc_cell = ws.cell(i, col)
        wc_cell.value = round(service_time[j], 3)
        j += 1

    wb.save(excel_path)


# calculate_arrival_time()
# calculate_inter_arrival_time()

def calculate_total_service_time():
    # clerk 1 service before testing
    calculate_service_time(200, 300, 500, 5)
    # clerk 2 service before testing
    calculate_service_time(200, 300, 500, 6)
    # lightning check
    calculate_service_time(35, 45, 55, 8)
    # break check
    calculate_service_time(30, 60, 90, 10)
    # steering check
    calculate_service_time(30, 60, 90, 12)
    # gas check
    calculate_service_time(60, 90, 120, 14)
    # clerk 1 service after testing
    calculate_service_time(80, 120, 190, 17)
    # clerk 2 service after testing
    calculate_service_time(80, 120, 190, 18)


calculate_total_service_time()

